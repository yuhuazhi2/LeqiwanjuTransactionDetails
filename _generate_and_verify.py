# -*- coding: utf-8 -*-
"""
独立运行：重新生成框架雏形 + 立即验证列结构
"""
import sys, os, logging
logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))

from src.database.connector import DatabaseConnector
from src.database.ufsystem import UFSystemQuerier

# ===== 1. 连接数据库，获取账套列表 =====
db_config = {
    "server": "server03", "port": 1433,
    "username": "sa", "password": "YUHUAZHIsql1",
    "timeout": 30, "charset": "GBK",
}

connector = DatabaseConnector(db_config)
connector.connect("UFSystem")
querier = UFSystemQuerier(connector)
all_accounts = querier.get_all_accounts()
accounts = [acc for acc in all_accounts if acc.cAcc_Id not in ("998", "999")]
year_map = querier.get_all_account_years_map()

print(f"查询到 {len(accounts)} 个有效账套")

# ===== 2. 构造 builder 并生成框架雏形 =====
from src.report.builder import ReportBuilder

config = {
    "database": db_config,
    "account_filter": {"include_ids": [], "exclude_ids": ["998", "999"]},
    "template": {
        "filepath": os.path.join(PROJECT_ROOT, "分店财务报表模板.xlsx"),
        "source_sheet": "sheet",
        "year_row": 1, "header_row": 2, "data_start_row": 3,
    },
    "output": {
        "dir": os.path.join(PROJECT_ROOT, "output"),
        "filename_prefix": "分店财务报表_",
        "file_extension": ".xlsx",
        "open_when_done": False,
    },
    "report_year": 2026,
    "report_months": [],
}

builder = ReportBuilder(config)
account_years = {acc.cAcc_Id: 2026 for acc in accounts}

print("\n===== 开始生成框架雏形 =====")
output_path = builder.build_framework(accounts=accounts, account_years=account_years)
print(f"生成完成: {output_path}")

# ===== 3. 立即验证列结构 =====
from openpyxl import load_workbook
print("\n===== 验证列结构 =====")
wb = load_workbook(output_path)

# 先看模板原始结构
template_path = os.path.join(PROJECT_ROOT, "分店财务报表模板.xlsx")
twb = load_workbook(template_path)
tws = twb[twb.sheetnames[0]]
template_headers = []
for col in range(1, tws.max_column + 1):
    v = tws.cell(2, col).value
    template_headers.append(str(v) if v else "")
print(f"原始模板 第2行: {template_headers}")
print(f"原始模板 列数: {len(template_headers)}")
twb.close()

# 看生成结果
for sname in wb.sheetnames:
    ws = wb[sname]
    headers = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(2, col).value
        headers.append(str(v) if v else "")
    print(f"\n[{sname}]")
    print(f"  第2行: {headers}")
    print(f"  列数: {len(headers)}")

wb.close()
connector.close()
print("\n===== 验证完成 =====")