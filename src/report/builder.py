"""
报表生成器
========
核心类：ReportBuilder
负责协调模板解析、数据提取和Excel生成全流程。
"""

import os
import logging
from datetime import datetime
from typing import Optional

from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from src.database.connector import DatabaseConnector
from src.database.ufsystem import UFSystemQuerier, AccountInfo
from src.database.t3data import T3DataExtractor
from src.template.parser import TemplateParser, TemplateLayout

logger = logging.getLogger(__name__)


class ReportBuilder:
    """
    报表生成器
    流程: 加载模板 → 查询账套列表 → 遍历账套取数 → 填充到各页签
    """

    # 默认样式
    HEADER_FONT = Font(name="微软雅黑", bold=True, size=11)
    HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                              fill_type="solid")
    DATA_FONT = Font(name="微软雅黑", size=10)
    TITLE_FONT = Font(name="微软雅黑", bold=True, size=14)
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    CURRENCY_FORMAT = '#,##0.00'

    # ================================================================
    # 行背景色配置（用于 build_framework 的 _apply_row_colors）
    # ================================================================
    # 各关键行的背景色（柔和的浅色系，便于长时间阅读）
    FILL_PERFORMANCE = PatternFill(start_color="E8D5F5", end_color="E8D5F5",
                                   fill_type="solid")       # 业绩完成率 - 浅紫色
    FILL_TARGET = PatternFill(start_color="D5F5E3", end_color="D5F5E3",
                              fill_type="solid")            # 目标预算 - 浅绿色
    FILL_SALES = PatternFill(start_color="D6EAF8", end_color="D6EAF8",
                             fill_type="solid")             # 销售业绩 - 浅蓝色
    FILL_COST = PatternFill(start_color="FDEBD0", end_color="FDEBD0",
                            fill_type="solid")              # 主营业务成本 - 浅橙色
    FILL_EXPENSE_HEADER = PatternFill(start_color="F9E79F", end_color="F9E79F",
                                      fill_type="solid")    # 营业/管理/财务费用（标题行）- 深黄色
    FILL_EXPENSE = PatternFill(start_color="FCF3CF", end_color="FCF3CF",
                               fill_type="solid")           # 营业/管理/财务费用明细行 - 浅黄色
    FILL_EXPENSE_ALT = PatternFill(start_color="FEF9E7", end_color="FEF9E7",
                                   fill_type="solid")       # 费用明细交替色 - 浅米色
    FILL_RATE = PatternFill(start_color="F2F3F4", end_color="F2F3F4",
                            fill_type="solid")              # 费用率 - 浅灰色
    FILL_PROFIT = PatternFill(start_color="AED6F1", end_color="AED6F1",
                              fill_type="solid")            # 利润 - 中浅蓝色
    FILL_DIVIDEND = PatternFill(start_color="F5B7B1", end_color="F5B7B1",
                                fill_type="solid")          # 分红 - 浅粉色
    FILL_BANK = PatternFill(start_color="D2B4DE", end_color="D2B4DE",
                            fill_type="solid")              # 银行余额 - 淡紫色

    # ================================================================
    # 模板行标签 → T3科目映射表
    # key: 模板中的行标签（标准化后），value: 数据提取方法
    # 这是核心映射配置，后续可扩展为外部配置文件
    # ================================================================
    LABEL_TO_DATA_KEY = {
        # 销售业绩（各渠道）
        "油菜花":     "油菜花收入",
        "现金":       "现金收入",
        "美团":       "美团收入",
        "抖音":       "抖音收入",
        "其他业务收入": "其他业务收入",
        # 成本
        "主营业务成本": "__cost__",
        # 营业费用
        "广告费":     "广告费",
        "物料费":     "物料费",
        "设备":       "设备",
        "折旧费":     "折旧费",
        "房租":       "房租",
        "物业费":     "物业费",
        "电费":       "电费",
        "修配费":     "修配费",
        "运杂费":     "运杂费",
        "其他":       "其他",
        # 管理费用
        "工资":       "工资",
        "办公费":     "办公费",
        "差旅费":     "差旅费",
        "业务招待费": "业务招待费",
        "员工福利":   "员工福利",
        "装修费":     "装修费",
        "开办费":     "开办费",
        "服务咨询费": "服务咨询费",
        "社保":       "社保",
        "管理公司费用分摊": "管理公司费用分摊",
        "奖金":       "奖金",
        "税费":       "税费",
        # 财务费用
        "手续费":     "手续费",
    }

    def __init__(self, config: dict):
        """
        :param config: 全局配置字典（从settings.yaml加载）
        """
        self.config = config
        self.db_config = config["database"]
        self.template_config = config["template"]
        self.output_config = config["output"]
        self.account_filter = config["account_filter"]
        self.report_year = config["report_year"]
        self.report_months = config.get("report_months", [])

        # 核心组件
        self.connector = DatabaseConnector(self.db_config)
        self.ufsystem = UFSystemQuerier(self.connector)
        self.extractor = T3DataExtractor(self.connector)
        self._template_parser: Optional[TemplateParser] = None
        self.template_path = self._resolve_template_path()

        # 输出
        self._wb: Optional[Workbook] = None

    @property
    def template_parser(self) -> TemplateParser:
        if self._template_parser is None:
            # 优先取配置中指定的 source_sheet，未指定则传空字符串，
            # TemplateParser 内部会自动取第一个sheet
            sheet_name = self.template_config.get("source_sheet", "")
            self._template_parser = TemplateParser(
                self.template_path,
                sheet_name
            )
        return self._template_parser

    def _resolve_template_path(self) -> str:
        """解析模板文件路径（支持相对/绝对路径）"""
        path = self.template_config["filepath"]
        if not os.path.isabs(path):
            # 相对于项目根目录
            root = os.path.dirname(os.path.dirname(
                os.path.dirname(__file__)))
            path = os.path.join(root, path)
        if not os.path.exists(path):
            raise FileNotFoundError(f"模板文件不存在: {path}")
        return path

    # ================================================================
    # 方法 build_framework — 生成报表初步框架雏形（V2.0 新增）
    # ================================================================
    # 功能说明：
    #   以"分店财务报表模板.xlsx"为雏形，从 UA_Account 表查询所有
    #   合规账套（cacc_name字段），每个账套复制一个表页，表页名称
    #   即为账套名称，在每个表页的 A2 位置用红色字体填入账套名称。
    #
    # 与现有 build() 方法的区别：
    #   - build()：完整流程（查询模板 → 取数 → 填充数据 → 出报表）
    #   - build_framework()：仅做框架（复制模板页签 + 写入账套名），
    #     不填充任何财务数据，供客户预览效果
    # ================================================================

    def build_framework(self, accounts: list = None,
                        account_years: dict[str, int] = None) -> str:
        """
        生成报表框架雏形（V2.0新增）
        ========================
        功能：以模板为基础，为每个合规账套复制一个表页，
             在A2位置用红色字体填入"账套名称+年度"，
             形成报表初步框架。
             生成的工作簿不保留模板原sheet，直接按账套名称命名各表页。

        流程（基于 openpyxl.copy_worksheet 在同一工作簿内复制）：
          1. 加载模板文件，获取模板sheet（取第一个sheet，或配置中指定的sheet）
          2. 如果未传入 accounts 参数，则从 UA_Account 表自动查询
          3. 对每个账套，用 copy_worksheet() 复制模板sheet，
             再以账套名重命名新sheet
          4. 所有账套复制完毕后，删除原始的模板sheet
          5. 在每个新sheet的 A2 单元格写入"账套名称+年度"并设为红色字体
          6. 保存到 output/ 目录

        参数:
            accounts: 指定的账套列表（可选），
                      不传则自动查询全部合规账套（过滤 998/999）
            account_years: 账套年份映射 {账套号: 年度}（可选），
                           指定每个账套对应的报表年份，用于A2显示。
                           未传入时，A2只显示账套名称。

        返回:
            str: 生成的Excel文件绝对路径
        """
        logger.info("=" * 60)
        logger.info("开始生成报表框架雏形")
        logger.info(f"模板: {self.template_path}")

        # ---- 步骤1：如果未传入账套列表，则自动从 UA_Account 查询 ----
        if accounts is None:
            accounts = self._get_accounts()
        logger.info(f"共查询到 {len(accounts)} 个合规账套")

        if not accounts:
            logger.warning("未查询到任何账套，请检查数据库配置！")
            raise ValueError("未查询到账套信息")

        # ---- 步骤2：加载模板工作簿 ----
        # 直接使用模板工作簿，在其内部用 copy_worksheet 复制sheet
        # 这是 openpyxl 原生支持的同一工作簿内复制方式，不会出现
        # 跨工作簿的 StyleProxy 兼容性问题
        self._wb = load_workbook(self.template_path)
        # 获取源sheet名称：优先取配置中指定的 sheet 名，
        # 如果配置未指定或找不到，则取模板文件中的第一个 sheet
        template_sheet_name = self.template_config.get("source_sheet", "")
        if not template_sheet_name or template_sheet_name not in self._wb.sheetnames:
            template_sheet_name = self._wb.sheetnames[0]
            logger.info(f"自动获取模板sheet名称: {template_sheet_name}")
        template_ws = self._wb[template_sheet_name]

        # ---- 步骤3：遍历每个账套，用 copy_worksheet 复制sheet ----
        for acc in accounts:
            sheet_name = acc.sheet_name[:31]  # Excel sheet名最多31字符
            logger.info(f"  处理账套: {acc.cAcc_Name} (ID={acc.cAcc_Id})")
            # copy_worksheet 在同一工作簿内复制全部内容和样式（原生支持）
            new_ws = self._wb.copy_worksheet(template_ws)
            new_ws.title = sheet_name

        # ---- 步骤4：删除原始的模板sheet，只保留账套sheet ----
        if template_sheet_name in self._wb.sheetnames:
            del self._wb[template_sheet_name]

        # ---- 步骤5：遍历每个账套sheet，在A2填入"账套名称+年度"并设红色字体 ----
        for acc in accounts:
            sheet_name = acc.sheet_name[:31]
            if sheet_name in self._wb.sheetnames:
                ws = self._wb[sheet_name]
                a2_cell = ws.cell(row=2, column=1)

                # 构建A2显示内容：账套名称 + 年度
                year_str = ""
                if account_years and acc.cAcc_Id in account_years:
                    year_str = str(account_years[acc.cAcc_Id])
                if year_str:
                    a2_cell.value = f"{acc.cAcc_Name} {year_str}"
                else:
                    a2_cell.value = acc.cAcc_Name

                a2_cell.font = Font(
                    name="微软雅黑", size=11, bold=True,
                    color="FF0000"  # 红色
                )

        # ---- 步骤5.5：根据各账套数据库已结账月份，动态调整列结构 ----
        for acc in accounts:
            sheet_name = acc.sheet_name[:31]
            if sheet_name not in self._wb.sheetnames:
                continue
            ws = self._wb[sheet_name]

            # 获取该账套对应的年度
            year_str = ""
            if account_years and acc.cAcc_Id in account_years:
                year_str = str(account_years[acc.cAcc_Id])
            if not year_str:
                continue  # 无年度信息，跳过列调整

            # 构建数据库名（重新计算，不依赖 acc.db_name 缓存的值）
            # 用友T3数据库命名规则：UFDATA_账套号_年度
            db_name = f"UFDATA_{acc.cAcc_Id.zfill(3)}_{year_str}"  # 如 UFDATA_007_2026
            try:
                closed_months = self.connector.get_closed_periods(db_name)
                if closed_months:
                    logger.info(f"  {sheet_name}: 已结账月份 {closed_months}，调整列结构")
                    self._adjust_sheet_columns(ws, closed_months)
                else:
                    logger.info(f"  {sheet_name}: 无已结账月份记录，保持默认列")
            except Exception as e:
                logger.warning(f"  {sheet_name} 列调整失败: {e}，使用默认列结构")

        # ---- 步骤5.6：根据各账套 code 表的 5101 科目，动态调整行结构 ----
        for acc in accounts:
            sheet_name = acc.sheet_name[:31]
            if sheet_name not in self._wb.sheetnames:
                continue
            ws = self._wb[sheet_name]

            # 获取该账套对应的年度
            year_str = ""
            if account_years and acc.cAcc_Id in account_years:
                year_str = str(account_years[acc.cAcc_Id])
            if not year_str:
                continue  # 无年度信息，跳过行调整

            # 构建数据库名
            db_name = f"UFDATA_{acc.cAcc_Id.zfill(3)}_{year_str}"
            try:
                self._adjust_sheet_subject_rows(ws, db_name)
            except Exception as e:
                logger.warning(f"  {sheet_name} 行调整失败: {e}，使用默认行结构")

            # ---- 步骤5.7：根据各账套 code 表的 5501 科目，动态调整营业费用明细行结构 ----
            try:
                self._adjust_sheet_expense_rows(ws, db_name)
            except Exception as e:
                logger.warning(f"  {sheet_name} 营业费用行调整失败: {e}，使用默认行结构")

            # ---- 步骤5.8：根据各账套 code 表的 5502 科目，动态调整管理费用与财务费用之间明细行结构 ----
            try:
                self._adjust_sheet_manage_rows(ws, db_name)
            except Exception as e:
                logger.warning(f"  {sheet_name} 管理费用明细行调整失败: {e}，使用默认行结构")

        # ---- 步骤5.9：为每个账套工作表设置行背景色（增强可读性） ----
        for acc in accounts:
            sheet_name = acc.sheet_name[:31]
            if sheet_name not in self._wb.sheetnames:
                continue
            ws = self._wb[sheet_name]
            try:
                self._apply_row_colors(ws)
            except Exception as e:
                logger.warning(f"  {sheet_name} 行背景色应用失败: {e}")

        # ---- 步骤5.10：从 GL_AccVouch 表填充期间损益结转数据（V2.6 新增） ----
        for acc in accounts:
            sheet_name = acc.sheet_name[:31]
            if sheet_name not in self._wb.sheetnames:
                continue
            ws = self._wb[sheet_name]

            # 获取该账套对应的年度
            year_str = ""
            if account_years and acc.cAcc_Id in account_years:
                year_str = str(account_years[acc.cAcc_Id])
            if not year_str:
                continue  # 无年度信息，跳过

            # 构建数据库名
            db_name = f"UFDATA_{acc.cAcc_Id.zfill(3)}_{year_str}"
            try:
                self._fill_period_transfer_data(ws, db_name)
            except Exception as e:
                logger.warning(f"  {sheet_name} 期间损益结转数据填列失败: {e}")

        # ---- 步骤5.11：计算各汇总行（销售业绩/营业费用/管理费用）的合计值（V2.8 新增） ----
        for acc in accounts:
            sheet_name = acc.sheet_name[:31]
            if sheet_name not in self._wb.sheetnames:
                continue
            ws = self._wb[sheet_name]
            try:
                self._calculate_summary_sums(ws)
            except Exception as e:
                logger.warning(f"  {sheet_name} 汇总行合计计算失败: {e}")

        # ---- 步骤5.12：从 GL_AccSum 表查询银行余额并填列到"银行余额"行（V3.0 新增） ----
        # 银行余额 = 银行存款（1002科目）的期末余额 me 字段
        # 需要从各账套对应的数据库 UFDATA_XXX_YYYY 的 GL_AccSum 表中，
        # 按 iperiod（月份）字段匹配月份列，取 me（期末余额）字段的值填入对应单元格。
        for acc in accounts:
            sheet_name = acc.sheet_name[:31]
            if sheet_name not in self._wb.sheetnames:
                continue
            ws = self._wb[sheet_name]

            # 获取该账套对应的年度
            year_str = ""
            if account_years and acc.cAcc_Id in account_years:
                year_str = str(account_years[acc.cAcc_Id])
            if not year_str:
                continue  # 无年度信息，跳过

            # 构建数据库名（用友T3数据库命名规则：UFDATA_账套号_年度）
            db_name = f"UFDATA_{acc.cAcc_Id.zfill(3)}_{year_str}"
            try:
                self._fill_bank_balance(ws, db_name, int(year_str))
            except Exception as e:
                logger.warning(f"  {sheet_name} 银行余额填列失败: {e}")

        # ---- 步骤5.13：为每个账套修复"业绩完成率"行的Excel公式（V2.9 新增） ----
        # 需要在所有行/列结构调整完毕之后执行，确保行列号已稳定
        for acc in accounts:
            sheet_name = acc.sheet_name[:31]
            if sheet_name not in self._wb.sheetnames:
                continue
            ws = self._wb[sheet_name]
            try:
                self._fix_performance_formula(ws)
            except Exception as e:
                logger.warning(f"  {sheet_name} 业绩完成率公式修复失败: {e}")

        # ---- 步骤6：调整所有表页的合计列宽度（增加50%） ----
        for sheet_name in self._wb.sheetnames:
            ws = self._wb[sheet_name]
            try:
                self._widen_total_column(ws)
            except Exception as e:
                logger.warning(f"  {sheet_name} 合计列宽度调整失败: {e}")

        # ---- 步骤7：保存工作簿到输出目录 ----
        output_path = self._save_framework_workbook()

        logger.info(f"框架雏形生成完成，共 {len(accounts)} 个表页")
        logger.info(f"文件路径: {output_path}")
        return output_path

    # ================================================================
    # _adjust_sheet_columns — 根据已结账月份调整工作表列结构（V2.1 新增）
    # ================================================================
    # 功能说明：
    #   根据 gl_mend 表查询到的已结账月份列表，动态调整工作表的月份列：
    #   - 已结账月份存在的列 → 保留
    #   - 已结账月份不存在的多余月份列 → 删除该列
    #   - 新增已结账月份但模板中没有的列 → 在"年度合计"列前插入
    #
    # 使用 openpyxl 原生 delete_cols / insert_cols 操作列，自动维护
    # 右侧所有单元格的偏移。
    # ================================================================

    def _adjust_sheet_columns(self, ws, closed_months: list[int]):
        """
        根据已结账月份列表，动态调整单张工作表的月份列结构。

        策略：
          1. 扫描第1行（列标题），建立 {月份: 列号} 映射，找到"年度合计"列
          2. 计算需要删除的多余月份列 和 需要新增的月份列
          3. 删除多余列（从右向左删除，避免列号偏移）
          4. 缺少的月份列在"年度合计"前插入

        :param ws: 目标工作表
        :param closed_months: 已结账月份列表（升序），如 [1,2,3,4,5]
        """
        month_map = {
            "1月": 1, "2月": 2, "3月": 3, "4月": 4,
            "5月": 5, "6月": 6, "7月": 7, "8月": 8,
            "9月": 9, "10月": 10, "11月": 11, "12月": 12,
        }

        # ---- 1. 分析当前列结构：找出月份列和年度合计列 ----
        # 模板的列标题在第2行（第1行为空）
        existing_months = {}  # {月份: 列号}
        total_col = None      # "年度合计"列号

        for col_idx in range(1, ws.max_column + 1):
            # 先检查第1行，如果为空则检查第2行
            cell_val = ws.cell(1, col_idx).value
            if cell_val is None:
                cell_val = ws.cell(2, col_idx).value
            cell_str = str(cell_val).strip() if cell_val else ""

            if "合计" in cell_str or "汇总" in cell_str:
                total_col = col_idx
            elif cell_str in month_map:
                month = month_map[cell_str]
                existing_months[month] = col_idx

        if total_col is None or not existing_months:
            logger.warning("  _adjust_sheet_columns: 无法定位月份列或合计列，跳过调整")
            return

        # ---- 2. 计算需要删除/新增的月份 ----
        target_months = sorted(set(closed_months))  # 已结账月份（升序）
        current_months = sorted(existing_months.keys())  # 当前月份（升序）

        months_to_remove = [m for m in current_months if m not in target_months]
        months_to_add = [m for m in target_months if m not in current_months]

        logger.debug(f"    当前月份列: {current_months}, 目标: {target_months}")
        logger.debug(f"    需删除列: {months_to_remove}, 需新增: {months_to_add}")

        # ---- 3. 删除多余月份列（从右向左删，避免偏移） ----
        for month in sorted(months_to_remove, reverse=True):
            col_idx = existing_months[month]
            logger.debug(f"    删除列: 第{col_idx}列 ({month}月)")
            ws.delete_cols(col_idx, 1)
            # 删除后更新 existing_months 和 total_col 的列号
            total_col = total_col - 1 if total_col > col_idx else total_col
            # 更新 existing_months 中所有在删除列右侧的列号
            for m, c in list(existing_months.items()):
                if c > col_idx:
                    existing_months[m] = c - 1

        # ---- 4. 在"年度合计"列前插入缺少的月份列 ----
        # 先确定目标月份列应该按什么顺序插入到合计列之前
        # 从合计列向左，需要补全从1月到12月中缺失的月份
        added_count = 0
        for month in sorted(months_to_add):
            # 在 total_col 位置插入空列（原有列向右移）
            insert_pos = total_col
            logger.debug(f"    插入列: 在第{insert_pos}列前 ({month}月)")
            ws.insert_cols(insert_pos, 1)
            # 设置新列标题为月份（模板列标题在第2行，所以写入row2）
            # 同时写入row1以保持一致性
            for row_num in (1, 2):
                new_cell = ws.cell(row_num, insert_pos)
                new_cell.value = f"{month}月" if row_num == 2 else None
                if row_num == 2:
                    new_cell.font = self.HEADER_FONT
                    new_cell.alignment = Alignment(horizontal="center")
                    new_cell.border = self.THIN_BORDER
            # ---- 复制数据行的背景色和边框：从左侧列（前一个月列）复制到新列 ----
            # insert_cols 将原有列右移，新列在 insert_pos，左侧列在 insert_pos - 1
            left_col = insert_pos - 1
            for row_num in range(3, ws.max_row + 1):
                left_cell = ws.cell(row_num, left_col)
                new_cell = ws.cell(row_num, insert_pos)
                # 复制填充（背景色）
                try:
                    if left_cell.fill and left_cell.fill.fill_type:
                        new_cell.fill = copy(left_cell.fill)
                except Exception:
                    pass
                # 复制边框
                try:
                    if left_cell.border:
                        new_cell.border = copy(left_cell.border)
                except Exception:
                    pass
            total_col += 1  # 合计列右移一位
            added_count += 1

        if months_to_remove or months_to_add:
            logger.info(f"    列调整完成: 删除{len(months_to_remove)}列, 新增{added_count}列")

    # ================================================================
    # _adjust_sheet_subject_rows — 根据 code 表的 5101 科目调整行结构（V2.2 新增）
    # ================================================================
    # 功能说明：
    #   模板中 A5=销售业绩，A10=主营业务成本。
    #   根据各账套 code 表中 ccode 开头为 5101 的 6 位数字科目列表，
    #   动态调整 A5 与 A10 之间的行：
    #   - 先删除 A5 与 A10 之间原有的全部行（原模板中的渠道明细行）
    #   - 再根据 5101 科目数量插入对应行数
    #   - 每行的 A 列填入对应科目的 ccode_name
    # ================================================================

    def _adjust_sheet_subject_rows(self, ws, db_name: str):
        """
        根据指定账套 code 表中 5101 开头的科目列表，
        动态调整工作表中 A5（销售业绩）与 A10（主营业务成本）之间的行。

        :param ws: 目标工作表
        :param db_name: 账套数据库名，如 UFDATA_007_2024
        """
        # ---- 1. 查询该账套的 5101 科目列表 ----
        revenue_subjects = self.extractor.get_revenue_subjects_from_code(db_name)
        target_count = len(revenue_subjects)

        # 固定行号（模板中 A5=销售业绩，A10=主营业务成本）
        REVENUE_ROW = 5          # A5：销售业绩
        COST_ROW = 10            # A10：主营业务成本

        # ---- 2. 删除 A5 与 A10 之间的现有全部行（行6~行9） ----
        rows_to_remove = list(range(REVENUE_ROW + 1, COST_ROW))  # [6,7,8,9]
        for row_idx in sorted(rows_to_remove, reverse=True):
            ws.delete_rows(row_idx, 1)
            logger.debug(f"    删除第 {row_idx} 行（调整成本明细区域）")

        # 删除行后，A10 主营业务成本上移，新位置变为 A6
        current_cost_row = REVENUE_ROW + 1  # 删除后主营业务成本在 A6

        # ---- 3. 插入 5101 科目行（在"主营业务成本"行之前插入） ----
        if target_count > 0:
            logger.info(
                f"    插入 {target_count} 行 5101 科目："
                f"{[s['ccode_name'] for s in revenue_subjects]}"
            )
            for i, subject in enumerate(revenue_subjects):
                # 每次在当前主营业务成本行位置插入新行
                insert_row = current_cost_row
                ws.insert_rows(insert_row, 1)

                # 在新行 A 列填入科目名称（前加缩进空格，与模板中渠道明细行风格一致）
                cell = ws.cell(insert_row, 1)
                cell.value = f" {subject['ccode_name']}"
                # 复制样式：从上一行 A 列复制字体、边框、对齐、填充
                prev_row = insert_row - 1
                if prev_row >= 1:
                    prev_cell = ws.cell(prev_row, 1)
                    try:
                        if prev_cell.font:
                            cell.font = copy(prev_cell.font)
                    except Exception:
                        pass
                    try:
                        if prev_cell.border:
                            cell.border = copy(prev_cell.border)
                    except Exception:
                        pass
                    try:
                        if prev_cell.alignment:
                            cell.alignment = copy(prev_cell.alignment)
                    except Exception:
                        pass
                    try:
                        if prev_cell.fill and prev_cell.fill.fill_type:
                            cell.fill = copy(prev_cell.fill)
                    except Exception:
                        pass

                # ---- 为新插入行的其余全部数据列（B列起）加上细线边框 ----
                # 先找到"年度合计"列位置，只应用到合计列为止，避免合计列右侧多出边框
                total_col = ws.max_column  # 默认上限
                for col in range(1, ws.max_column + 1):
                    for check_row in (1, 2):
                        cell_val = ws.cell(check_row, col).value
                        cell_str = str(cell_val).strip() if cell_val else ""
                        if "合计" in cell_str or "汇总" in cell_str:
                            total_col = col
                            break
                    else:
                        continue
                    break
                for col_idx in range(2, total_col + 1):
                    data_cell = ws.cell(insert_row, col_idx)
                    try:
                        data_cell.border = copy(self.THIN_BORDER)
                    except Exception:
                        pass

                current_cost_row += 1
        else:
            logger.info(f"    该账套无 5101 科目，不插入明细行")

    # ================================================================
    # _adjust_sheet_expense_rows — 根据 code 表的 5501 科目调整营业费用行结构（V2.3 新增）
    # ================================================================
    # 功能说明：
    #   模板中"营业费用"与"管理费用"之间的行，根据各账套 code 表中
    #   ccode 开头为 5501 的 6 位数字科目列表动态调整：
    #   - 先删除"营业费用"与"管理费用"之间原有的全部现有行
    #   - 再根据 5501 科目列表插入：每个科目一行（填科目名称），
    #     然后隔一空白行，再插下一个科目，如此交替
    # ================================================================

    def _adjust_sheet_expense_rows(self, ws, db_name: str):
        """
        根据指定账套 code 表中 5501 开头的科目列表，
        动态调整工作表中"营业费用"与"管理费用"之间的行。

        :param ws: 目标工作表
        :param db_name: 账套数据库名，如 UFDATA_007_2024
        """
        # ---- 1. 查询该账套的 5501 科目列表 ----
        expense_subjects = self.extractor.get_expense_subjects_from_code(db_name)
        if not expense_subjects:
            logger.info("    该账套无 5501 科目，不调整营业费用明细行")
            return
        target_count = len(expense_subjects)

        # ---- 2. 扫描 A 列（所有行），找出"营业费用"和"管理费用"的行号 ----
        expense_row = None   # "营业费用"行号
        manage_row = None    # "管理费用"行号
        for row_idx in range(1, ws.max_row + 1):
            cell_val = ws.cell(row_idx, 1).value
            cell_str = str(cell_val).strip() if cell_val else ""
            if cell_str == "营业费用":
                expense_row = row_idx
            elif cell_str == "管理费用":
                manage_row = row_idx
            if expense_row is not None and manage_row is not None:
                break

        if expense_row is None or manage_row is None:
            logger.warning("    无法定位'营业费用'或'管理费用'行，跳过营业费用明细行调整")
            return
        if manage_row - expense_row <= 1:
            logger.warning("    '营业费用'与'管理费用'之间无空间，跳过")
            return

        logger.info(f"    定位: 营业费用行{expense_row}, 管理费用行{manage_row}")

        # ---- 3. 删除两者之间的现有全部行 ----
        # 删除范围: expense_row+1 到 manage_row-1
        rows_to_remove = list(range(expense_row + 1, manage_row))
        for row_idx in sorted(rows_to_remove, reverse=True):
            ws.delete_rows(row_idx, 1)
            logger.debug(f"    删除第 {row_idx} 行（营业费用明细区域）")

        # 删除后，"管理费用"行上移
        current_expense_bottom = expense_row + 1  # 此时"管理费用"在这里

        # ---- 4. 查找合计列位置（用于加边框） ----
        total_col = ws.max_column
        for col in range(1, ws.max_column + 1):
            for check_row in (1, 2):
                cell_val = ws.cell(check_row, col).value
                cell_str = str(cell_val).strip() if cell_val else ""
                if "合计" in cell_str or "汇总" in cell_str:
                    total_col = col
                    break
            else:
                continue
            break

        # ---- 5. 插入 5501 科目行 + 空白行（交替） ----
        # 在每个科目行之后留一行空白
        logger.info(
            f"    插入 {target_count} 行 5501 科目（交替留空白行）："
            f"{[s['ccode_name'] for s in expense_subjects]}"
        )
        for i, subject in enumerate(expense_subjects):
            # 先插入科目名称行
            insert_row = current_expense_bottom
            ws.insert_rows(insert_row, 1)

            cell = ws.cell(insert_row, 1)
            cell.value = f" {subject['ccode_name']}"

            # 设置 A 列字体为宋体12号加粗（与费用率等标题行一致）
            cell.font = Font(name="宋体", size=12, bold=True)
            # A 列加边框（与数据区域统一）
            cell.border = copy(self.THIN_BORDER)

            # B 列到合计列加边框
            for col_idx in range(2, total_col + 1):
                try:
                    ws.cell(insert_row, col_idx).border = copy(self.THIN_BORDER)
                except Exception:
                    pass

            current_expense_bottom += 1  # 科目行插入后下移

            # 插入空白行（留一行备用）
            blank_row = current_expense_bottom
            ws.insert_rows(blank_row, 1)
            # 空白行A列也加边框
            try:
                ws.cell(blank_row, 1).border = copy(self.THIN_BORDER)
            except Exception:
                pass
            # 空白行数据列加边框
            for col_idx in range(2, total_col + 1):
                try:
                    ws.cell(blank_row, col_idx).border = copy(self.THIN_BORDER)
                except Exception:
                    pass
            current_expense_bottom += 1  # 空白行后继续下移

        logger.info(f"    营业费用明细行调整完成: 插入 {target_count} 个科目（各间隔一行空白）")

    # ================================================================
    # _adjust_sheet_manage_rows — 根据 code 表的 5502 科目调整管理费用与财务费用之间的行（V2.4 新增）
    # ================================================================
    # 功能说明：
    #   模板中"管理费用"与"财务费用"之间的行，根据各账套 code 表中
    #   ccode 开头为 5502 的 6 位数字科目列表动态调整：
    #   - 先删除"管理费用"与"财务费用"之间原有的全部现有行（模板中写死的固定明细）
    #   - 再根据 5502 科目列表插入：每个科目一行（填科目名称 + 缩进），
    #     然后隔一空白行，再插下一个科目，如此交替
    # ================================================================

    def _adjust_sheet_manage_rows(self, ws, db_name: str):
        """
        根据指定账套 code 表中 5502 开头的科目列表，
        动态调整工作表中"管理费用"与"财务费用"之间的行。

        :param ws: 目标工作表
        :param db_name: 账套数据库名，如 UFDATA_007_2024
        """
        # ---- 1. 查询该账套的 5502 科目列表 ----
        manage_subjects = self.extractor.get_manage_subjects_from_code(db_name)
        if not manage_subjects:
            logger.info("    该账套无 5502 科目，不调整管理费用明细行")
            return
        target_count = len(manage_subjects)

        # ---- 2. 扫描 A 列，找出"管理费用"和"财务费用"的行号 ----
        manage_row = None   # "管理费用"行号
        finance_row = None  # "财务费用"行号
        for row_idx in range(1, ws.max_row + 1):
            cell_val = ws.cell(row_idx, 1).value
            cell_str = str(cell_val).strip() if cell_val else ""
            if cell_str == "管理费用":
                manage_row = row_idx
            elif cell_str == "财务费用":
                finance_row = row_idx
            if manage_row is not None and finance_row is not None:
                break

        if manage_row is None or finance_row is None:
            logger.warning("    无法定位'管理费用'或'财务费用'行，跳过管理费用明细行调整")
            return
        if finance_row - manage_row <= 1:
            logger.warning("    '管理费用'与'财务费用'之间无空间，跳过")
            return

        logger.info(f"    定位: 管理费用行{manage_row}, 财务费用行{finance_row}")

        # ---- 3. 删除两者之间的现有全部行 ----
        rows_to_remove = list(range(manage_row + 1, finance_row))
        for row_idx in sorted(rows_to_remove, reverse=True):
            ws.delete_rows(row_idx, 1)
            logger.debug(f"    删除第 {row_idx} 行（管理费用明细区域）")

        # 删除后，"财务费用"行上移
        current_manage_bottom = manage_row + 1  # 此时"财务费用"在这里

        # ---- 4. 查找合计列位置（用于加边框） ----
        total_col = ws.max_column
        for col in range(1, ws.max_column + 1):
            for check_row in (1, 2):
                cell_val = ws.cell(check_row, col).value
                cell_str = str(cell_val).strip() if cell_val else ""
                if "合计" in cell_str or "汇总" in cell_str:
                    total_col = col
                    break
            else:
                continue
            break

        # ---- 5. 插入 5502 科目行 + 空白行（交替） ----
        logger.info(
            f"    插入 {target_count} 行 5502 科目（交替留空白行）："
            f"{[s['ccode_name'] for s in manage_subjects]}"
        )
        for i, subject in enumerate(manage_subjects):
            # 先插入科目名称行（带缩进空格）
            insert_row = current_manage_bottom
            ws.insert_rows(insert_row, 1)

            cell = ws.cell(insert_row, 1)
            cell.value = f" {subject['ccode_name']}"

            # 设置 A 列字体为宋体12号加粗（与费用率等标题行一致）
            cell.font = Font(name="宋体", size=12, bold=True)
            # A 列加边框（与数据区域统一）
            cell.border = copy(self.THIN_BORDER)

            # B 列到合计列加边框
            for col_idx in range(2, total_col + 1):
                try:
                    ws.cell(insert_row, col_idx).border = copy(self.THIN_BORDER)
                except Exception:
                    pass

            current_manage_bottom += 1  # 科目行插入后下移

            # 插入空白行（留一行备用）
            blank_row = current_manage_bottom
            ws.insert_rows(blank_row, 1)
            # 空白行A列也加边框
            try:
                ws.cell(blank_row, 1).border = copy(self.THIN_BORDER)
            except Exception:
                pass
            # 空白行数据列加边框
            for col_idx in range(2, total_col + 1):
                try:
                    ws.cell(blank_row, col_idx).border = copy(self.THIN_BORDER)
                except Exception:
                    pass
            current_manage_bottom += 1  # 空白行后继续下移

        logger.info(f"    管理费用明细行调整完成: 插入 {target_count} 个科目（各间隔一行空白）")

    # ================================================================
    # _apply_row_colors — 根据 A 列行标签为工作表各行设置背景色（V2.5 新增）
    # ================================================================
    # 功能说明：
    #   在行结构全部调整完毕后，扫描 A 列标签，为特定行分配不同的背景色
    #   以增强可读性。适用于 build_framework() 流程末尾。
    #
    # 颜色规则（按用户要求）：
    #   - "业绩完成率"               → 浅紫色     (FILL_PERFORMANCE)
    #   - "目标预算（一档）"         → 浅绿色     (FILL_TARGET)
    #   - "销售业绩"                 → 浅蓝色     (FILL_SALES)
    #   - "主营业务成本"             → 浅橙色     (FILL_COST)
    #   - "营业费用"/"管理费用"/"财务费用"（标题行）→ 深黄色 (FILL_EXPENSE_HEADER)
    #   - "营业费用"与"管理费用"之间的明细行 → 字符行用浅黄、空行保留原色
    #   - "管理费用"与"财务费用"之间的明细行 → 同上
    #   - "费用率"                   → 浅灰色     (FILL_RATE)
    #   - "利润"                     → 中浅蓝色   (FILL_PROFIT)
    #   - "分红"/"总分红"            → 浅粉色     (FILL_DIVIDEND)
    #   - "银行余额"                 → 淡紫色     (FILL_BANK)
    # ================================================================

    def _apply_row_colors(self, ws):
        """
        为工作表的各行（从A列到年度合计列）设置背景色。

        应该在行结构调整（_adjust_sheet_*_rows）全部完成后调用，
        确保行号已稳定。

        :param ws: 目标工作表
        """
        # ---- 1. 找到"年度合计"列位置，只应用到该列为止 ----
        total_col = ws.max_column
        for col in range(1, ws.max_column + 1):
            for check_row in (1, 2):
                cell_val = ws.cell(check_row, col).value
                cell_str = str(cell_val).strip() if cell_val else ""
                if "合计" in cell_str or "汇总" in cell_str:
                    total_col = col
                    break
            else:
                continue
            break

        # ---- 2. 建立精确行标签 → 填充色映射表 ----
        # 注意：A列单元格值可能含前导空格，用 strip() 匹配
        LABEL_FILL_MAP = {
            "业绩完成率":         self.FILL_PERFORMANCE,
            "目标预算（一档）":   self.FILL_TARGET,
            "销售业绩":           self.FILL_SALES,
            "主营业务成本":       self.FILL_COST,
            "费用率":             self.FILL_RATE,
            "利润":               self.FILL_PROFIT,
            "利润率":             self.FILL_PROFIT,
            "分红":               self.FILL_DIVIDEND,
            "总分红（往年分红万）": self.FILL_DIVIDEND,
            "投资":               self.FILL_DIVIDEND,
            "银行余额":           self.FILL_BANK,
        }

        # ---- 3. 先扫描 A 列，找出三个费用标题行的位置 ----
        expense_row = None   # "营业费用"行号
        manage_row = None    # "管理费用"行号
        finance_row = None   # "财务费用"行号

        for row_idx in range(1, ws.max_row + 1):
            cell_val = ws.cell(row_idx, 1).value
            cell_str = str(cell_val).strip() if cell_val else ""
            if cell_str == "营业费用":
                expense_row = row_idx
            elif cell_str == "管理费用":
                manage_row = row_idx
            elif cell_str == "财务费用":
                finance_row = row_idx
            # 三个都找到即可提前结束扫描
            if expense_row is not None and manage_row is not None and finance_row is not None:
                break

        logger.debug(f"  _apply_row_colors 定位: 营业费用行{expense_row}, "
                      f"管理费用行{manage_row}, 财务费用行{finance_row}")

        # ---- 4. 遍历 A 列，应用颜色 ----
        for row_idx in range(1, ws.max_row + 1):
            cell_val = ws.cell(row_idx, 1).value
            cell_str = str(cell_val).strip() if cell_val else ""
            if not cell_str:
                continue

            fill = None

            # 4a. 基础映射匹配
            if cell_str in LABEL_FILL_MAP:
                fill = LABEL_FILL_MAP[cell_str]

            # 4b. 三大费用标题行：用深黄色（与明细行的浅黄色区分）
            elif cell_str in ("营业费用", "管理费用", "财务费用"):
                fill = self.FILL_EXPENSE_HEADER

            # 如果找到填充色，应用到整行
            if fill is not None:
                self._apply_row_fill(ws, row_idx, total_col, fill)

        # ---- 5. 处理"营业费用"与"管理费用"之间的明细行（有内容的行设浅黄色，空行保留原色） ----
        if expense_row is not None and manage_row is not None:
            for row_idx in range(expense_row + 1, manage_row):
                cell_val = ws.cell(row_idx, 1).value
                cell_str = str(cell_val).strip() if cell_val else ""
                if cell_str:
                    # 有字符内容的行 → 浅黄色
                    self._apply_row_fill(ws, row_idx, total_col, self.FILL_EXPENSE)
                # 空白行不设背景色，保留原样

        # ---- 6. 处理"管理费用"与"财务费用"之间的明细行（有内容的行设浅黄色，空行保留原色） ----
        if manage_row is not None and finance_row is not None:
            for row_idx in range(manage_row + 1, finance_row):
                cell_val = ws.cell(row_idx, 1).value
                cell_str = str(cell_val).strip() if cell_val else ""
                if cell_str:
                    self._apply_row_fill(ws, row_idx, total_col, self.FILL_EXPENSE)
                # 空白行不设背景色，保留原样

    @staticmethod
    def _apply_row_fill(ws, row: int, total_col: int, fill: PatternFill):
        """
        给指定行的 A 列到 total_col 列设置背景填充色。

        :param ws: 工作表
        :param row: 行号
        :param total_col: 最后一列号（含）
        :param fill: 填充样式
        """
        for col_idx in range(1, total_col + 1):
            cell = ws.cell(row, col_idx)
            try:
                cell.fill = copy(fill)
            except Exception:
                pass

    # ================================================================
    # _fill_period_transfer_data — 从 GL_AccVouch 填列期间损益结转数据（V2.6 新增）
    # ================================================================
    # 功能说明：
    #   从 GL_AccVouch 表查询 cdigest='期间损益结转' 的全部记录，
    #   按以下规则将数据填入工作表的对应行列：
    #
    #   【收入类科目（5101xx）】→ 取 md（借方）
    #     例如 ccode=510101 对应 code 表中 ccode_name=油菜花，
    #     则填入"油菜花"这一行对应月份列。
    #
    #   【成本类科目（5401）】→ 取 mc（贷方）
    #   【费用类科目（5501xx/5502xx/5503）】→ 取 mc（贷方）
    #
    #   iperiod = 1 → 1月列，iperiod = 2 → 2月列，依此类推。
    #   注意：不再对 ccode_equal 做任何过滤，所有期间损益结转记录都参与填列。
    # ================================================================

    def _fill_period_transfer_data(self, ws, db_name: str):
        """
        从 GL_AccVouch 期间损益结转记录填列数据到页签。

        :param ws: 目标工作表
        :param db_name: 账套数据库名，如 UFDATA_007_2026
        """
        # ---- 1. 获取月份列映射：扫描表头（第1行或第2行）建立 {月份: 列号} ----
        month_map = {
            "1月": 1, "2月": 2, "3月": 3, "4月": 4,
            "5月": 5, "6月": 6, "7月": 7, "8月": 8,
            "9月": 9, "10月": 10, "11月": 11, "12月": 12,
        }
        month_col_map = {}  # {月份数字: Excel列号}
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(1, col_idx).value
            if cell_val is None:
                cell_val = ws.cell(2, col_idx).value
            cell_str = str(cell_val).strip() if cell_val else ""
            if cell_str in month_map:
                month_col_map[month_map[cell_str]] = col_idx

        if not month_col_map:
            logger.warning("  _fill_period_transfer_data: 无法定位月份列，跳过")
            return
        logger.debug(f"  月份列映射: {month_col_map}")

        # ---- 2. 查询该账套的期间损益结转凭证 ----
        # 从 db_name 中提取年度（UFDATA_XXX_YYYY）
        year = None
        try:
            parts = db_name.split('_')
            if len(parts) >= 3:
                year = int(parts[-1])
        except (ValueError, IndexError):
            pass

        if not year:
            logger.warning(f"  _fill_period_transfer_data: 无法从 {db_name} 提取年度，跳过")
            return

        vouchers = self.extractor.get_period_transfer_vouchers(db_name)
        if not vouchers:
            logger.info(f"  {db_name}: 无期间损益结转凭证数据，跳过")
            return

        logger.info(f"  {db_name}: 获取 {len(vouchers)} 条期间损益结转分录")

        # ---- 3. 收集所有涉及的 ccode，批量查询科目名称 ----
        all_ccodes = list(set(v["ccode"] for v in vouchers))
        code_name_map = self.extractor.get_code_subject_name_batch(db_name, all_ccodes)
        logger.debug(f"  科目名称映射: {code_name_map}")

        # ---- 4. 扫描 A 列，建立 {科目名称: 行号} 映射 ----
        label_row_map = {}  # {ccode_name: row_index}
        for row_idx in range(1, ws.max_row + 1):
            cell_val = ws.cell(row_idx, 1).value
            if cell_val:
                label = str(cell_val).strip()
                if label:
                    label_row_map[label] = row_idx

        # ---- 5. 遍历每条凭证分录，填列到对应行列 ----
        # 不再对 ccode_equal 做任何过滤（取消 ccocode_equal=3131 限制），
        # 所有期间损益结转记录（收入类取md借方，成本费用类取mc贷方）都参与填列。
        filled_count = 0
        for vouch in vouchers:
            ccode = str(vouch.get("ccode", "")).strip()
            iperiod = int(vouch.get("iperiod", 0))
            md_val = float(vouch.get("md", 0) or 0)
            mc_val = float(vouch.get("mc", 0) or 0)

            # 查找该 ccode 对应的科目名称
            ccode_name = code_name_map.get(ccode, "")
            if not ccode_name:
                continue

            # 查找科目名称对应的行号
            row_idx = label_row_map.get(ccode_name, None)
            if row_idx is None:
                continue

            # 查找月份对应的列号
            col_idx = month_col_map.get(iperiod, None)
            if col_idx is None:
                continue

            # ---- 确定填入 md 还是 mc ----
            # 收入类（5101xx）→ md（借方）
            # 成本/费用类（5401/5501/5502/5503）→ mc（贷方）
            is_revenue = ccode.startswith("5101")
            is_cost = ccode.startswith("5401")
            is_expense = (ccode.startswith("5501") or
                          ccode.startswith("5502") or
                          ccode.startswith("5503"))

            if is_revenue:
                val = md_val
            elif is_cost or is_expense:
                val = mc_val
            else:
                # 其他科目跳过（不处理）
                continue

            # ---- 跳过零值 ----
            if val == 0:
                continue

            # ---- 写入单元格 ----
            cell = ws.cell(row_idx, col_idx)
            # 如果单元格已有值，累加（同一科目同一月份可能有多个分录）
            existing = cell.value
            try:
                existing = float(existing) if existing else 0
            except (ValueError, TypeError):
                existing = 0
            cell.value = existing + val
            # 设置数字格式
            cell.number_format = self.CURRENCY_FORMAT
            cell.font = self.DATA_FONT
            cell.alignment = Alignment(horizontal="right")
            cell.border = self.THIN_BORDER

            filled_count += 1
            logger.debug(f"    填入: {ccode_name} ({ccode}), iperiod={iperiod}, "
                         f"col={col_idx}, row={row_idx}, val={val}")

        logger.info(f"  {db_name}: 期间损益结转数据填列完成，共{filled_count}个单元格")

    # ================================================================
    # _calculate_summary_sums — 计算汇总行（销售业绩/营业费用/管理费用）的合计值（V2.8 新增）
    # ================================================================
    # 功能说明：
    #   在期间损益结转数据填列完毕后，根据明细行数值计算并填入汇总行的合计值。
    #
    #   计算规则：
    #   - 销售业绩行：取"销售业绩"与"主营业务成本"之间所有行（每列）数值合计
    #   - 营业费用行：取"营业费用"与"管理费用"之间 有背景色 的行（每列）数值合计
    #   - 管理费用行：取"管理费用"与"财务费用"之间 有背景色 的行（每列）数值合计
    #
    #   备注：
    #   - "有背景色的行"即内容不为空且已设置过背景填充色的明细科目行；
    #     空行（无背景色）不参与营业费用和管理费用的合计计算。
    #   - 销售业绩合计不区分空白行，取区间内所有行的合计（区间内不含空白行）。
    # ================================================================

    # ---- 辅助常量：月份标题映射 ----
    _MONTH_MAP = {
        "1月": 1, "2月": 2, "3月": 3, "4月": 4,
        "5月": 5, "6月": 6, "7月": 7, "8月": 8,
        "9月": 9, "10月": 10, "11月": 11, "12月": 12,
    }

    def _calculate_summary_sums(self, ws):
        """
        计算各汇总行的合计值并填入对应单元格。
        应在 _fill_period_transfer_data 和 _apply_row_colors 之后调用。

        计算策略（区分列处理）：
          1. 先遍历各月份列，将区间内有效明细行的值合计后填入汇总行对应月份单元格
          2. 再遍历合计列，将其值设为汇总行自身各月份单元格之和
             （而非从明细行直接汇总合计列）

        :param ws: 目标工作表
        """
        # ---- 1. 扫描 A 列，定位关键行 ----
        sales_row = None       # "销售业绩"行号
        cost_row = None        # "主营业务成本"行号
        expense_row = None     # "营业费用"行号
        manage_row = None      # "管理费用"行号
        finance_row = None     # "财务费用"行号
        profit_row = None      # "利润"行号
        profit_rate_row = None # "利润率"行号

        for row_idx in range(1, ws.max_row + 1):
            cell_val = ws.cell(row_idx, 1).value
            cell_str = str(cell_val).strip() if cell_val else ""
            if cell_str == "销售业绩":
                sales_row = row_idx
            elif cell_str == "主营业务成本":
                cost_row = row_idx
            elif cell_str == "营业费用":
                expense_row = row_idx
            elif cell_str == "管理费用":
                manage_row = row_idx
            elif cell_str == "财务费用":
                finance_row = row_idx
            elif cell_str == "利润":
                profit_row = row_idx
            elif cell_str == "利润率":
                profit_rate_row = row_idx

        logger.debug(f"  _calculate_summary_sums 定位: "
                      f"销售业绩行{sales_row}, 主营业务成本行{cost_row}, "
                      f"营业费用行{expense_row}, 管理费用行{manage_row}, "
                      f"财务费用行{finance_row}, "
                      f"利润行{profit_row}, 利润率行{profit_rate_row}")

        # ---- 2. 查找月份列和合计列 ----
        month_col_map = {}  # {月份数字: 列号}
        total_col = None
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(1, col_idx).value
            if cell_val is None:
                cell_val = ws.cell(2, col_idx).value
            cell_str = str(cell_val).strip() if cell_val else ""
            if cell_str in self._MONTH_MAP:
                month_col_map[self._MONTH_MAP[cell_str]] = col_idx
            elif "合计" in cell_str or "汇总" in cell_str:
                total_col = col_idx

        month_cols = list(month_col_map.values())  # 各月份列号
        if not month_cols:
            logger.warning("  _calculate_summary_sums: 未找到月份列，跳过")
            return

        logger.debug(f"  月份列: {month_col_map}, 合计列: {total_col}")

        # ---- 工具函数：写入汇总行单元格，并设置格式 ----
        def _write_summary_cell(ws_row, ws_col, value):
            cell = ws.cell(ws_row, ws_col)
            cell.value = value
            cell.number_format = self.CURRENCY_FORMAT
            cell.font = self.DATA_FONT
            cell.alignment = Alignment(horizontal="right")
            cell.border = self.THIN_BORDER

        # ---- 工具函数：安全读取单元格数值 ----
        def _get_cell_float(ws_row, ws_col):
            cell_val = ws.cell(ws_row, ws_col).value
            try:
                return float(cell_val) if cell_val else 0.0
            except (ValueError, TypeError):
                return 0.0

        # ======== 3. 计算各汇总行（仅月份列） ========
        # 逐列处理，而不是逐行处理，避免重复行列扫描

        # --- 3a. 销售业绩：取销售业绩与主营业务成本之间所有行 ---
        if sales_row is not None and cost_row is not None and sales_row < cost_row:
            logger.debug(f"  计算销售业绩合计: 区间行 {sales_row+1} ~ {cost_row-1}")
            for col_idx in month_cols:
                total = 0.0
                for row_idx in range(sales_row + 1, cost_row):
                    cell_val = ws.cell(row_idx, col_idx).value
                    try:
                        total += float(cell_val) if cell_val else 0
                    except (ValueError, TypeError):
                        pass
                _write_summary_cell(sales_row, col_idx, total)

        # --- 3b. 营业费用：仅统计有背景色的明细行 ---
        if expense_row is not None and manage_row is not None and expense_row < manage_row:
            logger.debug(f"  计算营业费用合计: 区间行 {expense_row+1} ~ {manage_row-1}")
            for col_idx in month_cols:
                total = 0.0
                for row_idx in range(expense_row + 1, manage_row):
                    cell_a = ws.cell(row_idx, 1)
                    has_fill = (cell_a.fill and cell_a.fill.fill_type is not None
                                and cell_a.fill.fill_type != "")
                    if not has_fill:
                        continue
                    cell_val = ws.cell(row_idx, col_idx).value
                    try:
                        total += float(cell_val) if cell_val else 0
                    except (ValueError, TypeError):
                        pass
                _write_summary_cell(expense_row, col_idx, total)

        # --- 3c. 管理费用：仅统计有背景色的明细行 ---
        if manage_row is not None and finance_row is not None and manage_row < finance_row:
            logger.debug(f"  计算管理费用合计: 区间行 {manage_row+1} ~ {finance_row-1}")
            for col_idx in month_cols:
                total = 0.0
                for row_idx in range(manage_row + 1, finance_row):
                    cell_a = ws.cell(row_idx, 1)
                    has_fill = (cell_a.fill and cell_a.fill.fill_type is not None
                                and cell_a.fill.fill_type != "")
                    if not has_fill:
                        continue
                    cell_val = ws.cell(row_idx, col_idx).value
                    try:
                        total += float(cell_val) if cell_val else 0
                    except (ValueError, TypeError):
                        pass
                _write_summary_cell(manage_row, col_idx, total)

        # ======== 4. 计算所有行的合计列 ========
        # 遍历每一行（数据从第3行开始），对每个存在月份数据的行，
        # 计算合计列 = 本行各月份数值之和
        if total_col is not None:
            for row_idx in range(3, ws.max_row + 1):
                # 跳过利润和利润率行（合计列将由公式计算覆盖）
                if row_idx == profit_row or row_idx == profit_rate_row:
                    continue
                # 检查该行是否有月份数据（任一月份列非空且非零）
                has_month_data = False
                row_total = 0.0
                for col_idx in month_cols:
                    cell_val = ws.cell(row_idx, col_idx).value
                    try:
                        val = float(cell_val) if cell_val else 0
                    except (ValueError, TypeError):
                        val = 0
                    if val != 0:
                        has_month_data = True
                    row_total += val

                if has_month_data:
                    _write_summary_cell(row_idx, total_col, row_total)

        # ======== 5. 计算利润和利润率（按列公式计算） ========
        # 利润 = 销售业绩 - 主营业务成本 - 营业费用 - 管理费用 - 财务费用
        # 利润率 = 利润 / 销售业绩（百分比）
        # 对每个月份列和合计列分别计算

        # --- 5a. 计算各月份列的利润值 ---
        if profit_row is not None:
            all_cols = list(month_cols)
            if total_col is not None:
                all_cols.append(total_col)  # 合计列也按公式计算
            for col_idx in all_cols:
                sales_val = _get_cell_float(sales_row, col_idx) if sales_row else 0
                cost_val = _get_cell_float(cost_row, col_idx) if cost_row else 0
                expense_val = _get_cell_float(expense_row, col_idx) if expense_row else 0
                manage_val = _get_cell_float(manage_row, col_idx) if manage_row else 0
                finance_val = _get_cell_float(finance_row, col_idx) if finance_row else 0
                profit_val = sales_val - cost_val - expense_val - manage_val - finance_val
                _write_summary_cell(profit_row, col_idx, profit_val)

        # --- 5b. 计算各月份列的利润率 ---
        if profit_rate_row is not None:
            all_cols = list(month_cols)
            if total_col is not None:
                all_cols.append(total_col)
            for col_idx in all_cols:
                sales_val = _get_cell_float(sales_row, col_idx) if sales_row else 0
                profit_val = _get_cell_float(profit_row, col_idx) if profit_row else 0
                rate_val = (profit_val / sales_val * 100) if sales_val != 0 else 0
                # 写入百分比格式
                cell = ws.cell(profit_rate_row, col_idx)
                cell.value = rate_val
                cell.number_format = '0.00"%"'
                cell.font = self.DATA_FONT
                cell.alignment = Alignment(horizontal="right")
                cell.border = self.THIN_BORDER

        logger.info(f"  汇总行合计及利润/利润率计算完成")

    def build(self) -> str:
        """
        执行完整的报表生成流程
        :return: 生成的Excel文件路径
        """
        logger.info("=" * 60)
        logger.info("开始生成财务报表")
        logger.info(f"年份: {self.report_year}")
        logger.info(f"模板: {self.template_path}")

        # 1. 解析模板布局
        template = self._parse_template()
        logger.info(f"模板解析完成: {len(template.row_labels)}行, "
                    f"{len(template.columns)}列")

        # 2. 查询所有账套
        accounts = self._get_accounts()
        logger.info(f"共查询到 {len(accounts)} 个账套")

        if not accounts:
            logger.warning("未查询到任何账套，请检查数据库配置！")
            raise ValueError("未查询到账套信息")

        # 3. 创建工作簿并填充
        self._create_workbook(template, accounts)

        # 3.5 调整所有页签的合计列宽度（增加50%）
        for sheet_name in self._wb.sheetnames:
            ws = self._wb[sheet_name]
            try:
                self._widen_total_column(ws)
            except Exception as e:
                logger.warning(f"  {sheet_name} 合计列宽度调整失败: {e}")

        output_path = self._save_workbook()

        # 4. 可选：自动打开文件
        if self.output_config.get("open_when_done", True):
            self._open_file(output_path)

        logger.info(f"报表生成完成: {output_path}")
        return output_path

    def _parse_template(self) -> TemplateLayout:
        """解析模板文件"""
        return self.template_parser.parse()

    def _get_accounts(self) -> list[AccountInfo]:
        """获取需要生成报表的账套列表"""
        include = self.account_filter.get("include_ids", [])
        exclude = self.account_filter.get("exclude_ids", [])

        if include:
            accounts = []
            for aid in include:
                acc = self.ufsystem.get_account_by_id(aid)
                if acc:
                    accounts.append(acc)
                else:
                    logger.warning(f"账套 {aid} 未找到，已跳过")
            return accounts
        else:
            return self.ufsystem.get_filtered_accounts(exclude_ids=exclude)

    def _create_workbook(self, template: TemplateLayout,
                         accounts: list[AccountInfo]):
        """
        创建工作簿，为每个账套创建一个页签
        """
        # 加载模板作为基础样式参考
        template_wb = load_workbook(self.template_path)
        template_ws = template_wb[template.sheet_name]

        self._wb = Workbook()
        # 删除默认页
        self._wb.remove(self._wb.active)

        for acc in accounts:
            sheet_name = acc.sheet_name
            logger.info(f"  处理账套: {acc.cAcc_Name} ({acc.cAcc_Id})")

            # 构建数据库名（含年度）
            db_name = f"{acc.db_name}_{self.report_year}"

            # 创建新页签
            ws = self._wb.create_sheet(title=sheet_name)

            # 从模板复制结构和样式
            self._copy_sheet_structure(template_ws, ws, template)

            # 填充数据
            self._fill_sheet_data(ws, template, db_name)

        template_wb.close()
        logger.info(f"所有页签创建完成，共 {len(accounts)} 个")

    def _copy_sheet_structure(self, src_ws, dst_ws, template: TemplateLayout):
        """
        从模板页签复制结构和样式到目标页签
        （保留格式、行标签、表头，清空数据区域）
        """
        # 复制A列（行标签）和格式
        for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row,
                                    min_col=1, max_col=1):
            for cell in row:
                new_cell = dst_ws.cell(row=cell.row, column=1)
                new_cell.value = cell.value
                if cell.has_style:
                    # 按原始属性值重新构造 StyleProxy 对象，避免跨工作簿直接赋值
                    try:
                        new_cell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            color=cell.font.color,
                        )
                    except Exception:
                        pass
                    try:
                        new_cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            wrap_text=cell.alignment.wrap_text,
                        )
                    except Exception:
                        pass
                    try:
                        new_cell.fill = PatternFill(
                            fill_type=cell.fill.fill_type,
                            start_color=cell.fill.start_color,
                            end_color=cell.fill.end_color,
                        )
                    except Exception:
                        pass

        # 复制表头行（第2行）
        for col_idx, col_header in enumerate(template.columns, 1):
            new_cell = dst_ws.cell(row=1, column=col_idx)
            new_cell.value = col_header.label
            new_cell.font = self.HEADER_FONT
            new_cell.fill = self.HEADER_FILL
            new_cell.alignment = Alignment(horizontal="center")
            new_cell.border = self.THIN_BORDER

        # 设置店名（第一行第一列）
        dst_ws.cell(1, 1).value = template.store_name
        dst_ws.cell(1, 1).font = self.TITLE_FONT

    # ================================================================
    # 数据填充核心逻辑
    # ================================================================
    def _fill_sheet_data(self, ws, template: TemplateLayout, db_name: str):
        """
        向指定页签填充数据
        :param ws: 目标工作表
        :param template: 模板布局
        :param db_name: 账套数据库名，如 UFDATA_001_2024
        """
        months = self.report_months or list(range(1, 13))

        try:
            # 批量提取各科目数据
            revenue = self.extractor.get_monthly_revenue(
                db_name, self.report_year, months
            )
            cost = self.extractor.get_monthly_cost(
                db_name, self.report_year, months
            )
            expenses = self.extractor.get_monthly_expenses(
                db_name, self.report_year, months
            )

            # 合并所有数据源，方便查找
            all_data = {}
            all_data.update(revenue)        # 各渠道收入
            all_data.update(cost)           # 成本
            all_data.update(expenses)       # 各项费用

            # 遍历模板的每一行，匹配并填充数据
            for row_label in template.row_labels:
                row = row_label.row_index
                label_text = self._normalize_label(row_label.label)

                # ---- 查找数据源 ----
                data_key = self._find_matching_key(label_text)
                data_source = all_data.get(data_key)

                if data_source is None and data_key == "__cost__":
                    data_source = cost.get("主营业务成本", {})

                if data_source is None:
                    continue  # 无对应数据源的跳过（如合计行、空白行）

                # ---- 填充各月数据 ----
                for col_header in template.columns:
                    if col_header.month > 0:
                        val = data_source.get(col_header.month, 0)
                        self._set_cell_value(ws, row, col_header.col_index, val)
                    elif col_header.is_total:
                        total = sum(
                            data_source.get(m, 0) for m in months
                        )
                        self._set_cell_value(ws, row, col_header.col_index, total)

            # ---- 补充计算行（如毛利率、费用率等） ----
            self._fill_calculated_rows(ws, template, db_name, months)

            logger.debug(f"  {db_name} 数据填充完成")

        except Exception as e:
            logger.error(f"  {db_name} 数据提取失败: {e}", exc_info=True)
            ws.cell(2, 1).value = f"[数据提取错误] {e}"

    def _normalize_label(self, label: str) -> str:
        """标准化标签文字，去除空格和括号变体"""
        return (label.replace(" ", "")
                     .replace("（", "(")
                     .replace("）", ")")
                     .replace("　", "")
                     .strip())

    def _find_matching_key(self, label_text: str) -> Optional[str]:
        """
        根据标准化后的标签文字，在映射表中查找匹配的数据键
        使用包含匹配（子串匹配），以适应 "（1）油菜花" 这种带前缀的写法
        """
        # 精确匹配
        if label_text in self.LABEL_TO_DATA_KEY:
            return self.LABEL_TO_DATA_KEY[label_text]

        # 模糊匹配：查找映射表的key是否在标签文字中
        for map_key, data_key in self.LABEL_TO_DATA_KEY.items():
            if map_key in label_text or label_text in map_key:
                return data_key

        return None

    def _fill_calculated_rows(self, ws, template: TemplateLayout,
                               db_name: str, months: list[int]):
        """
        填充需要计算的行，如毛利、毛利率、费用率、利润等
        """
        # 获取各月总收入、总成本
        monthly_total_revenue = {}
        monthly_total_cost = {}
        monthly_total_expense = {}

        revenue = self.extractor.get_monthly_revenue(db_name, self.report_year, months)
        cost_data = self.extractor.get_monthly_cost(db_name, self.report_year, months)
        expense_data = self.extractor.get_monthly_expenses(db_name, self.report_year, months)

        # 汇总收入
        for _, mdata in revenue.items():
            for m, v in mdata.items():
                monthly_total_revenue[m] = monthly_total_revenue.get(m, 0) + v

        # 汇总成本
        cost = cost_data.get("主营业务成本", {})
        for m, v in cost.items():
            monthly_total_cost[m] = monthly_total_cost.get(m, 0) + v

        # 汇总费用
        for _, mdata in expense_data.items():
            for m, v in mdata.items():
                monthly_total_expense[m] = monthly_total_expense.get(m, 0) + v

        # 遍历模板行，找到需要计算的行
        for row_label in template.row_labels:
            label_text = row_label.label.strip()
            row = row_label.row_index

            if "毛利" == label_text and "毛利率" not in label_text:
                # 毛利 = 总收入 - 总成本
                for col in template.columns:
                    if col.month > 0:
                        val = (monthly_total_revenue.get(col.month, 0)
                               - monthly_total_cost.get(col.month, 0))
                        self._set_cell_value(ws, row, col.col_index, val)
                    elif col.is_total:
                        total_rev = sum(monthly_total_revenue.values())
                        total_cost = sum(monthly_total_cost.values())
                        self._set_cell_value(ws, row, col.col_index,
                                             total_rev - total_cost)

            elif "毛利率" == label_text:
                for col in template.columns:
                    if col.month > 0:
                        rev = monthly_total_revenue.get(col.month, 0)
                        cst = monthly_total_cost.get(col.month, 0)
                        val = (rev - cst) / rev * 100 if rev != 0 else 0
                        self._set_cell_value(ws, row, col.col_index, val)
                        # 毛利率显示为百分比
                        ws.cell(row, col.col_index).number_format = '0.00"%"'
                    elif col.is_total:
                        total_rev = sum(monthly_total_revenue.values())
                        total_cost = sum(monthly_total_cost.values())
                        val = (total_rev - total_cost) / total_rev * 100 if total_rev != 0 else 0
                        self._set_cell_value(ws, row, col.col_index, val)
                        ws.cell(row, col.col_index).number_format = '0.00"%"'

            elif "利润" == label_text and "利润率" not in label_text and "总分红" not in label_text:
                # 利润 = 毛利 - 费用
                for col in template.columns:
                    if col.month > 0:
                        gross = (monthly_total_revenue.get(col.month, 0)
                                 - monthly_total_cost.get(col.month, 0))
                        val = gross - monthly_total_expense.get(col.month, 0)
                        self._set_cell_value(ws, row, col.col_index, val)
                    elif col.is_total:
                        total_gross = (sum(monthly_total_revenue.values())
                                       - sum(monthly_total_cost.values()))
                        total_exp = sum(monthly_total_expense.values())
                        self._set_cell_value(ws, row, col.col_index,
                                             total_gross - total_exp)

            elif "利润率" == label_text:
                for col in template.columns:
                    if col.month > 0:
                        rev = monthly_total_revenue.get(col.month, 0)
                        cst = monthly_total_cost.get(col.month, 0)
                        exp = monthly_total_expense.get(col.month, 0)
                        profit = (rev - cst - exp)
                        val = profit / rev * 100 if rev != 0 else 0
                        self._set_cell_value(ws, row, col.col_index, val)
                        ws.cell(row, col.col_index).number_format = '0.00"%"'

    # ================================================================
    # _widen_total_column — 将"年度合计"列宽度增加50%（V2.7 新增）
    # ================================================================
    # 功能说明：
    #   扫描工作表第1行或第2行，找到"年度合计"（或包含"合计"/"汇总"）的列，
    #   将其列宽增加50%，使得财务数字显示更完整，避免显示为"####"。
    #   适用于 build_framework() 和 build() 两个流程。
    # ================================================================

    @staticmethod
    def _widen_total_column(ws):
        """
        将工作表中"年度合计"列的宽度增加50%。

        :param ws: 目标工作表
        """
        from openpyxl.utils import get_column_letter

        total_col = None
        # 扫描第1行和第2行，查找"年度合计"列
        for row_num in (1, 2):
            for col_idx in range(1, ws.max_column + 1):
                cell_val = ws.cell(row_num, col_idx).value
                cell_str = str(cell_val).strip() if cell_val else ""
                if "合计" in cell_str or "汇总" in cell_str:
                    total_col = col_idx
                    break
            if total_col is not None:
                break

        if total_col is None:
            logger.warning("  _widen_total_column: 无法定位合计列，跳过")
            return

        col_letter = get_column_letter(total_col)
        current_width = ws.column_dimensions[col_letter].width
        if current_width is None or current_width < 1:
            current_width = 10  # 默认宽度
            logger.debug(f"  {col_letter}列当前宽度为None，使用默认值10")
        new_width = round(current_width * 1.5, 1)  # 增加50%
        ws.column_dimensions[col_letter].width = new_width
        logger.info(f"  调整{col_letter}列（合计列）宽度: {current_width} → {new_width}")

    def _set_cell_value(self, ws, row: int, col: int, value: float):
        """设置单元格数值并应用格式"""
        cell = ws.cell(row=row, column=col)
        cell.value = value if value != 0 else 0
        cell.font = self.DATA_FONT
        cell.number_format = self.CURRENCY_FORMAT
        cell.alignment = Alignment(horizontal="right")
        cell.border = self.THIN_BORDER

    def _save_workbook(self) -> str:
        """保存（完整报表）工作簿到输出文件"""
        output_dir = self.output_config.get("dir", "output")
        os.makedirs(output_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        prefix = self.output_config.get("filename_prefix", "分店财务报表_")
        ext = self.output_config.get("file_extension", ".xlsx")
        filename = f"{prefix}{self.report_year}_{timestamp}{ext}"
        output_path = os.path.join(output_dir, filename)

        self._wb.save(output_path)
        return os.path.abspath(output_path)

    # ================================================================
    # _save_framework_workbook — 保存框架雏形工作簿（V2.0 新增）
    # ================================================================
    # 专用于 build_framework() 的保存方法，与 _save_workbook() 的区别：
    #   - 文件名加 "_框架雏形" 后缀，与完整报表区分
    #   - 其余逻辑相同（输出目录、时间戳、扩展名）
    # ================================================================

    def _save_framework_workbook(self) -> str:
        """
        保存框架雏形工作簿到输出文件
        ==============================
        生成的文件名格式如：分店财务报表_框架雏形_2024_20260626_112233.xlsx
        保存后自动关闭工作簿释放资源。
        """
        output_dir = self.output_config.get("dir", "output")
        os.makedirs(output_dir, exist_ok=True)

        # 生成带"框架雏形"后缀的文件名，便于与完整报表区分
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        prefix = self.output_config.get("filename_prefix", "分店财务报表_")
        ext = self.output_config.get("file_extension", ".xlsx")
        filename = f"{prefix}框架雏形_{self.report_year}_{timestamp}{ext}"
        output_path = os.path.join(output_dir, filename)

        self._wb.save(output_path)
        logger.info(f"框架雏形文件已保存: {output_path}")
        return os.path.abspath(output_path)

    # ================================================================
    # _fix_performance_formula — 修复"业绩完成率"行公式（V2.9 新增）
    # ================================================================
    # 功能说明：
    #   扫描当前工作表，找到"业绩完成率"行、"目标预算（一档）"行和"销售业绩"行，
    #   为业绩完成率行的每个月份列和年度合计列写入公式：
    #     =IF(<销售业绩单元格>=0,"",<目标预算单元格>/<销售业绩单元格>)
    #
    #   此方法应在所有行/列结构调整完成后调用，确保行列号已稳定。
    #   这样后续使用者手工填列第5行"销售业绩"时，
    #   第三行"业绩完成率"会自动计算更新。
    # ================================================================
    def _fix_performance_formula(self, ws):
        """
        为当前工作表的"业绩完成率"行写入Excel公式。

        :param ws: 目标工作表
        """
        # ---- 1. 扫描A列，定位关键行 ----
        rate_row = None     # "业绩完成率"行号
        target_row = None   # "目标预算（一档）"行号
        sales_row = None    # "销售业绩"行号

        for row_idx in range(1, ws.max_row + 1):
            cell_val = ws.cell(row_idx, 1).value
            cell_str = str(cell_val).strip() if cell_val else ""
            if cell_str == "业绩完成率":
                rate_row = row_idx
            elif cell_str == "目标预算（一档）" or cell_str == "目标预算":
                target_row = row_idx
            elif cell_str == "销售业绩":
                sales_row = row_idx
            if rate_row is not None and target_row is not None and sales_row is not None:
                break

        if rate_row is None:
            logger.warning("  _fix_performance_formula: 未找到'业绩完成率'行，跳过")
            return
        if target_row is None:
            logger.warning("  _fix_performance_formula: 未找到'目标预算'行，跳过")
            return
        if sales_row is None:
            logger.warning("  _fix_performance_formula: 未找到'销售业绩'行，跳过")
            return

        logger.debug(f"  _fix_performance_formula 定位: "
                      f"业绩完成率行{rate_row}, 目标预算行{target_row}, 销售业绩行{sales_row}")

        # ---- 2. 扫描月份列和合计列，为每个数据列写入公式 ----
        from openpyxl.utils import get_column_letter

        COLUMN_MAP = {
            "1月": 1, "2月": 2, "3月": 3, "4月": 4,
            "5月": 5, "6月": 6, "7月": 7, "8月": 8,
            "9月": 9, "10月": 10, "11月": 11, "12月": 12,
        }

        formula_count = 0
        for col_idx in range(2, ws.max_column + 1):  # B列起
            # 从第1行或第2行获取列标题
            col_label = None
            for check_row in (1, 2):
                cell_val = ws.cell(check_row, col_idx).value
                if cell_val:
                    col_label = str(cell_val).strip()
                    break
            if not col_label:
                continue

            # 判断是否为月份列或合计列
            is_month = col_label in COLUMN_MAP
            is_total = "合计" in col_label or "汇总" in col_label
            if not (is_month or is_total):
                continue

            col_letter = get_column_letter(col_idx)
            # 公式：=IF(F5=0,"",F4/F5) 对应模板原始逻辑
            # 其中F列替换为当前列，4替换为目标预算行号，5替换为销售业绩行号
            formula = (f'=IF({col_letter}{sales_row}=0,"",'
                       f'{col_letter}{target_row}/{col_letter}{sales_row})')

            cell = ws.cell(rate_row, col_idx)
            cell.value = formula
            # 设置数字格式为金额格式（因为结果是比率，用百分比格式更合适，但保留原模板风格）
            cell.number_format = self.CURRENCY_FORMAT
            cell.font = self.DATA_FONT
            cell.alignment = Alignment(horizontal="right")
            cell.border = self.THIN_BORDER
            formula_count += 1

        logger.info(f"  _fix_performance_formula: 已写入 {formula_count} 个公式单元格")

    # ================================================================
    # _fill_bank_balance — 从 GL_AccSum 总账表填列银行余额数据（V3.0 新增）
    # ================================================================
    # 功能说明：
    #   银行余额 = 银行存款（1002科目）的期末余额 me 字段。
    #
    #   流程：
    #     1. 扫描工作表第1行（或第2行）的列标题，建立 {月份: Excel列号} 映射
    #     2. 扫描 A 列，找到"银行余额"所在的行号
    #     3. 调用 T3DataExtractor.get_bank_balance_from_gl_accsum() 查询各月
    #        期末余额
    #     4. 将查询结果填入对应行列单元格
    #
    #   注意：
    #     - GL_AccSum 表的 iperiod 字段：1=1月，2=2月，……，12=12月
    #     - me（期末余额）字段直接取值，无额外计算
    #     - 如果某月份查无记录，该单元格留空（不写入0值，保持模板原始风格）
    # ================================================================

    def _fill_bank_balance(self, ws, db_name: str, year: int):
        """
        从 GL_AccSum 总账表查询银行存款（1002科目）的期末余额 me，
        填入工作表的"银行余额"行。

        银行余额 = 银行存款（1002科目）的期末余额 me 字段。

        :param ws: 目标工作表（openpyxl Worksheet 对象）
        :param db_name: 账套数据库名，如 UFDATA_007_2026
        :param year: 会计年度
        """
        # ---- 1. 建立月份列映射 {月份数字: Excel列号} ----
        # 月份标题映射
        MONTH_MAP = {
            "1月": 1, "2月": 2, "3月": 3,
            "4月": 4, "5月": 5, "6月": 6,
            "7月": 7, "8月": 8, "9月": 9,
            "10月": 10, "11月": 11, "12月": 12,
        }
        month_col_map = {}  # {月份数字: Excel列号}
        total_col = None    # "年度合计"列号

        for col_idx in range(1, ws.max_column + 1):
            # 先检查第1行，如果为空则检查第2行
            cell_val = ws.cell(1, col_idx).value
            if cell_val is None:
                cell_val = ws.cell(2, col_idx).value
            cell_str = str(cell_val).strip() if cell_val else ""

            if cell_str in MONTH_MAP:
                month_col_map[MONTH_MAP[cell_str]] = col_idx
            elif "合计" in cell_str or "汇总" in cell_str:
                total_col = col_idx

        if not month_col_map:
            logger.warning(f"  _fill_bank_balance: 无法定位月份列，跳过数据库 {db_name}")
            return
        logger.debug(f"  _fill_bank_balance 月份列映射: {month_col_map}, 合计列: {total_col}")

        # ---- 2. 扫描 A 列，找到"银行余额"所在行 ----
        bank_balance_row = None
        for row_idx in range(1, ws.max_row + 1):
            cell_val = ws.cell(row_idx, 1).value
            cell_str = str(cell_val).strip() if cell_val else ""
            if cell_str == "银行余额":
                bank_balance_row = row_idx
                break

        if bank_balance_row is None:
            logger.warning(f"  _fill_bank_balance: 未找到'银行余额'行，跳过数据库 {db_name}")
            return
        logger.debug(f"  _fill_bank_balance: 找到'银行余额'行号={bank_balance_row}")

        # ---- 3. 调用数据提取器查询各月银行余额 ----
        months = sorted(month_col_map.keys())  # 需要查询的月份列表
        balances = self.extractor.get_bank_balance_from_gl_accsum(
            db_name, year, months
        )

        if not balances:
            logger.info(f"  {db_name}: 银行余额查询无数据，跳过填列")
            return

        logger.info(f"  {db_name}: 银行余额查询结果: {balances}")

        # ---- 4. 将查询结果填入对应行列单元格 ----
        filled_count = 0
        for month_num, col_idx in month_col_map.items():
            balance = balances.get(month_num)
            if balance is None:
                # 该月份无数据，跳过（不写入0值）
                continue

            cell = ws.cell(bank_balance_row, col_idx)
            cell.value = balance
            cell.number_format = self.CURRENCY_FORMAT
            cell.font = self.DATA_FONT
            cell.alignment = Alignment(horizontal="right")
            cell.border = self.THIN_BORDER
            filled_count += 1
            logger.debug(f"    填列: 银行余额行{bank_balance_row}, "
                         f"{month_num}月(列{col_idx}), 余额={balance:.2f}")

        # ---- 5. 注意：银行余额的"年度合计"指标无实际意义，不填列 ----
        # 每月的银行余额相互独立，汇总合计无法反映任何经营信息，
        # 因此银行余额行与年度合计列的交叉单元格保持为空，不写入任何值。
        # 如果该单元格之前被复制过边框或格式，也不做额外清除处理，
        # 保持与模板原始风格一致。

        logger.info(f"  {db_name}: 银行余额填列完成，共 {filled_count} 个单元格")

    @staticmethod
    def _open_file(filepath: str):
        """尝试自动打开文件（Windows系统）"""
        try:
            os.startfile(filepath)
        except Exception as e:
            logger.warning(f"无法自动打开文件: {e}")