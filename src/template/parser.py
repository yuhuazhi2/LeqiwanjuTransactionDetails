"""
Excel模板解析器
==============
读取"分店财务报表模板.xlsx"，提取行列标签结构，
为数据填充提供行列映射关系。
"""

from dataclasses import dataclass, field
from typing import Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class RowLabel:
    """模板中的行标签"""
    row_index: int          # Excel行号（1-indexed）
    label: str              # 标签文字，如"销售业绩"
    level: int = 0          # 缩进层级（0=顶级, 1=二级）
    parent: Optional[str] = None  # 父级标签
    skip: bool = False      # 是否为汇总计算行（非直接取数）


@dataclass
class ColumnHeader:
    """模板中的列标题"""
    col_index: int          # Excel列号（1-indexed）
    label: str              # 标题文字，如"1月"
    is_total: bool = False  # 是否为"年度合计"列
    month: int = 0          # 月份数字（1-12），0表示合计列


@dataclass
class TemplateLayout:
    """模板布局信息"""
    sheet_name: str                 # 模板页签名
    row_labels: list[RowLabel]      # 行标签列表
    columns: list[ColumnHeader]     # 列标题列表
    months: list[int]               # 有效月份 [1..12]
    store_name: str = ""            # 店名
    data_start_row: int = 3         # 数据起始行
    data_start_col: int = 2         # 数据起始列（B列）
    column_count: int = 0           # 总列数


class TemplateParser:
    """模板解析器：读取模板文件的结构布局"""

    DEFAULT_ROW_TAGS = [
        "业绩完成率", "目标预算", "销售业绩",
        "主营业务成本", "毛利", "毛利率",
        "费用支出", "费用率", "营业费用",
        "管理费用", "财务费用",
        "利润", "利润率", "分红",
        "银行余额"
    ]

    def __init__(self, filepath: str, sheet_name: str = "三江店报表"):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self._wb = None

    def parse(self) -> TemplateLayout:
        """解析模板，返回布局信息"""
        self._wb = load_workbook(self.filepath, data_only=True)
        ws = self._wb[self.sheet_name]

        layout = TemplateLayout(
            sheet_name=self.sheet_name,
            row_labels=self._parse_rows(ws),
            columns=self._parse_columns(ws),
            months=[],
            data_start_row=ws.max_row + 1,
            data_start_col=self._find_data_start_col(ws),
            column_count=ws.max_column
        )

        # 提取月份信息
        for col in layout.columns:
            if col.month > 0:
                layout.months.append(col.month)

        # 提取店名（第一行第一列）
        layout.store_name = str(ws.cell(1, 1).value or "").strip()

        # 确定数据起始行
        for i, label in enumerate(layout.row_labels):
            if label.label in self.DEFAULT_ROW_TAGS:
                layout.data_start_row = label.row_index
                break

        self._wb.close()
        self._wb = None
        return layout

    def _parse_columns(self, ws: Worksheet) -> list[ColumnHeader]:
        """
        解析列标题（通常在第2行）
        格式: B1=店名, C1~N1=月份, O1=年度合计
        """
        columns = []
        month_map = {
            "1月": 1, "2月": 2, "3月": 3, "4月": 4,
            "5月": 5, "6月": 6, "7月": 7, "8月": 8,
            "9月": 9, "10月": 10, "11月": 11, "12月": 12,
        }

        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(1, col_idx).value
            cell_str = str(cell_val).strip() if cell_val else ""

            # 跳过空列
            if not cell_str:
                continue

            # 判断是否为合计列
            is_total = "合计" in cell_str or "汇总" in cell_str

            # 判断月份
            month = 0
            if not is_total:
                month = month_map.get(cell_str, 0)

            columns.append(ColumnHeader(
                col_index=col_idx,
                label=cell_str,
                is_total=is_total,
                month=month
            ))

        return columns

    def _parse_rows(self, ws: Worksheet) -> list[RowLabel]:
        """
        解析行标签（A列）
        逐行读取，根据缩进判断层级
        """
        row_labels = []
        max_row = ws.max_row

        for row_idx in range(1, max_row + 1):
            cell_val = ws.cell(row_idx, 1).value
            cell_str = str(cell_val).strip() if cell_val else ""

            # 跳过空行和页脚（编制/审核/签批）
            if not cell_str:
                continue
            if cell_str in ("编制：", "审核：", "签批："):
                break

            # 判断层级：以前导空格数判断
            raw = str(cell_val) if cell_val else ""
            leading_spaces = len(raw) - len(raw.lstrip())
            level = 0 if leading_spaces < 2 else 1

            row_labels.append(RowLabel(
                row_index=row_idx,
                label=cell_str.strip(),
                level=level,
                skip=self._is_calculated_row(cell_str.strip())
            ))

        return row_labels

    def _is_calculated_row(self, label: str) -> bool:
        """判断是否为计算行（非直接取数行）"""
        calculated = ["毛利", "毛利率", "利润", "利润率",
                      "费用支出", "费用率",
                      "银行余额", "总分红"]
        return label in calculated

    def _find_data_start_col(self, ws: Worksheet) -> int:
        """找到数据起始列（跳过A列的标签列）"""
        return 2  # B列通常是第一个数据列

    def row_label_to_key(self, label: str) -> str:
        """将行标签转换为标准化键名（用于配置映射匹配）"""
        return (label.replace("（", "(")
                     .replace("）", ")")
                     .replace(" ", "")
                     .strip())